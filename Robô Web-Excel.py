""" 
Robô que entra em paginas da Web, baixa arquivos em um diretório especifico e preenche uma planilha no excel
com esses dados.

            Criado por: Isabelly Cristine Lopes

            Você pode me encontrar em: 

            Linkedin ->  https://www.linkedin.com/in/isabelly-cristine-lopes-8a9b59204/
            Instagram -> @isabellyloppess
"""

# Imports

from selenium import webdriver
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import os
import openpyxl as xl
from datetime import datetime


# Opções do Navegador

options = webdriver.EdgeOptions()
options.add_argument("start-maximized")
options.add_argument("--disable-extensions")
prefs = {'download.default_directory': r'Onde você quer baixar o arquivo'}
options.add_experimental_option('prefs', prefs)
navegador = webdriver.Edge(EdgeChromiumDriverManager().install(), options=options)


# Funções

def find_css_send(value, keys):
        navegador.find_element(By.CSS_SELECTOR, value).send_keys(keys)
        
def find_xpath_click(value):
        navegador.find_element(By.XPATH, value).click()

def wdw_clickable_xpath(tempo, path):
        WebDriverWait(navegador, tempo).until(EC.element_to_be_clickable((By.XPATH, path))).click()

def wdw_clickable_id(tempo, path):
        WebDriverWait(navegador, tempo).until(EC.element_to_be_clickable((By.ID, path))).click()


def find_id_click(value):
        navegador.find_element(By.ID, value).click()


def preenche(caminho):
    wb2 = xl.load_workbook(caminho)
    ws2 = wb2.active
    mr = ws2.max_row
    mc = ws2.max_column

    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = ws2.cell(row=i, column=j)
            ws1.cell(row=i, column=j).value = c.value


def senha():
        WebDriverWait(navegador, 60).until(EC.element_to_be_clickable((By.ID, 'usuario'))).send_keys('***')
        WebDriverWait(navegador, 60).until(EC.element_to_be_clickable((By.ID, 'senha'))).send_keys('***')
        WebDriverWait(navegador, 60).until(EC.element_to_be_clickable((By.ID, 'empresa'))).send_keys('***')
        navegador.find_element(by=By.XPATH, value='/html/body/div[1]/div/form/button').click()


def menu_inicial():
        navegador.find_element(by=By.XPATH, value='//*[@id="side-menu"]/li[4]/a/span').click()
        navegador.find_element(by=By.XPATH, value='//*[@id="tipoRelatorio"]').click()

# Entrando no primeiro Site

navegador.get('SITE QUE DESEJA ACESSAR')

senha()

menu_inicial()

find_xpath_click('//*[@id="tipoRelatorio"]/option[2]')

find_id_click('dataIni')

sleep(2)

find_xpath_click('/html/body/div[5]/div[1]/table/thead/tr[1]/th[1]')


# Varrendo as datas do calendário do inicio do mês

todas_dates = navegador.find_elements(By.XPATH, "//div[@class='datepicker-days']//td")

def date_picker():
    for dataelemento in todas_dates:
        data = dataelemento.text
        if data == '1':
            dataelemento.click()
            break

date_picker()

find_id_click('dataFim')

sleep(2)

find_xpath_click('/html/body/div[5]/div[1]/table/thead/tr[1]/th[1]')


# Varrendo as datas do calendário do fim do mês

todas_dates_fim = navegador.find_elements(By.XPATH, "//div[@class='datepicker-days']//td[@class='day']")

for dataelemento in todas_dates_fim:
    data = dataelemento.text
    if data == '31':
         dataelemento.click()
    elif data == '30':
        dataelemento.click()
        break

# Configurando o arquivo a ser baixado 

find_xpath_click('//*[@id="cli"]')
find_xpath_click('//*[@id="cli"]/option[2]')

find_xpath_click('//*[@id="user"]')
find_xpath_click('//*[@id="user"]/option[2]')

find_xpath_click('//*[@id="produto"]')
find_xpath_click('//*[@id="produto"]/option[2]')

find_xpath_click('//*[@id="divTiposRelatorios"]/div/div[5]/div/button')
find_xpath_click('//*[@id="tipoconsulta"]/option[2]')

find_xpath_click('//*[@id="divTiposRelatorios"]/div/div[5]/div/button')


# Espera o botão de Exportar aparecer

sleep(20)
navegador.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
sleep(30)

wdw_clickable_xpath(60, '//*[@id="divBtnExportar"]/div/button/i')

WebDriverWait(navegador, 60).until(EC.element_to_be_clickable((By.CLASS_NAME, 'confirm'))).click()


# Entrando no segundo site

navegador.get('SITE QUE DESEJA ENTRAR')

senha()

navegador.find_element(By.ID, 'link-fila').click()

WebDriverWait(navegador, 60).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tabela_wrapper"]'
                                                                             '/div[1]/div/a[1]'))).click()

sleep(20)

navegador.quit()

# RENOMEANDO
data = datetime.today().strftime('%m%Y')

path = r'Onde seu arquivo foi baixado'
lista = os.listdir(path)

for i in lista:
        if 'RELATORIO_CONSUMO' in i:
                os.rename(path + i, path + f'BASE_RELATORIOCONSUMO{data}.xlsx')



# Interação com Excel

wb1 = xl.load_workbook('Arquivo que vai receber os dados.xlsx')

ws1 = wb1.worksheets[1]
preenche(os.path.join(path + f'BASE_RELATORIOCONSUMO{data}.xlsx'))


wb1.save(str('Arquivo que vai receber os dados.xlsx'))


 
